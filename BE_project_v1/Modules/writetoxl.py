from openpyxl import Workbook
from openpyxl import load_workbook


def write_ratingdata(path, ratings_data):
    wb = load_workbook(path)
    ws_wh = wb['WH']
    ws_fib = wb['FIB']
    row_wh = ws_wh.max_row + 1
    row_fib = ws_fib.max_row + 1

    count = ratings_data['count']
    wh_questions = ratings_data['wh_questions']
    cloze_questions = ratings_data['cloze_questions']
    gramRate = ratings_data['gramRate']
    ansRate = ratings_data['ansRate'] 
    diffRate = ratings_data['diffRate']
    conRate = ratings_data['conRate']
    filename = ratings_data['filename']
    timestamp = ratings_data['curr_datetime']

    #LOOP HERE
    j = 0
    for i in range(row_fib, row_fib+count):
        if j>=count:
            break
        ws_fib.cell(row=i,column=1).value = timestamp
        ws_fib.cell(row=i,column=2).value = filename
        ws_fib.cell(row=i,column=3).value = cloze_questions[j]['question']
        ws_fib.cell(row=i,column=4).value = cloze_questions[j]['answer']
        ws_fib.cell(row=i,column=5).value = ', '.join(cloze_questions[j]['distractors'])
        ws_fib.cell(row=i,column=6).value = float(gramRate[j])
        ws_fib.cell(row=i,column=7).value = float(ansRate[j])
        ws_fib.cell(row=i,column=8).value = float(diffRate[j])
        ws_fib.cell(row=i,column=9).value = float(conRate[j])
        j += 1
    #LOOP END

    
    #LOOP HERE
    j = 0
    for i in range(row_wh, row_wh+count):
        if j>=count:
            break
        ws_wh.cell(row=i,column=1).value = timestamp
        ws_wh.cell(row=i,column=2).value = filename
        ws_wh.cell(row=i,column=3).value = wh_questions[j][0]
        ws_wh.cell(row=i,column=4).value = wh_questions[j][1]
        ws_wh.cell(row=i,column=5).value = float(gramRate[count+j])
        ws_wh.cell(row=i,column=6).value = float(ansRate[count+j])
        ws_wh.cell(row=i,column=7).value = float(diffRate[count+j])
        ws_wh.cell(row=i,column=8).value = float(conRate[count+j])
        j += 1
    #LOOP END


    
    wb.save(path)



def download_data(data):
    wb = Workbook()
    ws_fib = wb.create_sheet('FIB')
    ws_wh = wb.create_sheet('WH')

    count = data['count']
    wh_questions = data['wh_questions']
    cloze_questions = data['cloze_questions']
    filename = data['filename']
    timestamp = data['curr_datetime']

    ws_fib.cell(row=1,column=1).value = "TIMESTAMP"
    ws_fib.cell(row=1,column=2).value = "INPUT_FILE"
    ws_fib.cell(row=1,column=3).value = "FIB_QUESTION"
    ws_fib.cell(row=1,column=4).value = "FIB_ANSWER"
    ws_fib.cell(row=1,column=5).value = "OPTIONS"

    ws_wh.cell(row=1,column=1).value = "TIMESTAMP"
    ws_wh.cell(row=1,column=2).value = "INPUT_FILE"
    ws_wh.cell(row=1,column=3).value = "WH_SENTENCE"
    ws_wh.cell(row=1,column=4).value = "WH_QUESTION"

    

    #LOOP HERE
    j = 0
    for i in range(2, 2+count):
        if j>=count:
            break
        ws_fib.cell(row=i,column=1).value = timestamp
        ws_fib.cell(row=i,column=2).value = filename
        ws_fib.cell(row=i,column=3).value = cloze_questions[j]['question']
        ws_fib.cell(row=i,column=4).value = cloze_questions[j]['answer']
        ws_fib.cell(row=i,column=5).value = ', '.join(cloze_questions[j]['distractors'])
        j += 1
    #LOOP END

    
    #LOOP HERE
    j = 0
    for i in range(2, 2+count):
        if j>=count:
            break
        ws_wh.cell(row=i,column=1).value = timestamp
        ws_wh.cell(row=i,column=2).value = filename
        ws_wh.cell(row=i,column=3).value = wh_questions[j][0]
        ws_wh.cell(row=i,column=4).value = wh_questions[j][1]
        j += 1
    #LOOP END


    filename = filename[-5::-1]
    filename = filename[::-1]
    wb.save("D:\\BE_Project\\BE_project_v1\\Downloads\\"+filename+".xlsx")
    return "D:\BE_Project\BE_project_v1\Downloads\\" + filename+ '.xlsx'
#write_ratingdata("D:\\BE_Project\\BE_project_v1\\Ratings.xlsx")
